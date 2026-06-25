"""
多模态附件抽取器
每个抽取器负责一种附件类型，将非结构化附件转为文本。
抽取器采用可插拔设计：未配置对应能力时优雅降级，记录占位说明。
"""

import csv
import io
import logging
from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def _get_attachment_payload(att: Dict[str, Any]) -> Optional[bytes]:
    """从附件 dict 中提取二进制内容（优先 data > path > url）"""
    import base64
    import os

    from core.config import get_settings

    # base64 内联数据
    data = att.get("data")
    if data:
        try:
            if isinstance(data, str):
                return base64.b64decode(data)
            if isinstance(data, (bytes, bytearray)):
                return bytes(data)
        except Exception as e:
            logger.warning("附件 base64 解码失败: %s", e)
            return None

    # 本地路径（防止路径遍历：仅允许在 attachment_dir 白名单目录内读取）
    path = att.get("path")
    if path:
        allowed_dir = os.path.realpath(get_settings().attachment_dir)
        real_path = os.path.realpath(path)
        if not real_path.startswith(allowed_dir + os.sep):
            logger.warning("附件路径越权访问被拒绝: %s (允许目录: %s)", path, allowed_dir)
            return None
        if os.path.exists(real_path):
            try:
                with open(real_path, "rb") as f:
                    return f.read()
            except Exception as e:
                logger.warning("附件文件读取失败 %s: %s", real_path, e)
                return None

    # URL（仅记录，不在抽取器内下载，避免阻塞；由上层预处理）
    url = att.get("url")
    if url:
        return None
    return None


def _detect_kind(mime: str, filename: str) -> str:
    """根据 mime / 文件名推断附件类型：text / image / audio / table / unknown"""
    mime = (mime or "").lower()
    fname = (filename or "").lower()

    # 表格类优先判断（text/csv 也属于表格，避免被 text/* 吞掉）
    if fname.endswith((".csv", ".tsv")) or mime in ("text/csv", "text/tab-separated-values"):
        return "table"
    if fname.endswith((".xlsx", ".xls")) or mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return "table"
    if fname.endswith(".pdf") or mime == "application/pdf":
        return "pdf"
    if mime.startswith("text/") or fname.endswith((".txt", ".md", ".log")):
        return "text"
    if mime.startswith("image/") or fname.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
        return "image"
    if mime.startswith("audio/") or fname.endswith((".wav", ".mp3", ".m4a", ".flac")):
        return "audio"
    if mime.startswith("video/") or fname.endswith((".mp4", ".mov")):
        return "video"
    return "unknown"


class BaseExtractor(ABC):
    """抽取器基类"""

    kind: str = "unknown"

    @abstractmethod
    async def extract(self, attachment: Dict[str, Any]) -> str:
        """抽取文本，返回纯文本"""
        raise NotImplementedError


class TextExtractor(BaseExtractor):
    """文本附件抽取器：直接读取内容"""

    kind = "text"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知文件")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            # 可能直接在 content 字段提供文本
            content = attachment.get("content")
            if content:
                return f"[文本附件 {filename}]\n{content}"
            return f"[文本附件 {filename}] 无法读取内容"
        try:
            text = payload.decode("utf-8", errors="replace")
            return f"[文本附件 {filename}]\n{text}"
        except Exception as e:
            logger.warning("文本附件解码失败 %s: %s", filename, e)
            return f"[文本附件 {filename}] 解码失败"


class TableExtractor(BaseExtractor):
    """表格附件抽取器：解析 CSV/TSV 为 Markdown 表格文本"""

    kind = "table"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知表格")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[表格附件 {filename}] 无法读取内容"

        fname = filename.lower()
        delimiter = "\t" if fname.endswith(".tsv") else ","

        # xlsx 需 openpyxl，未安装时降级
        if fname.endswith((".xlsx", ".xls")):
            try:
                return self._extract_xlsx(payload, filename)
            except ImportError:
                return f"[表格附件 {filename}] xlsx 解析需安装 openpyxl，已跳过"
            except Exception as e:
                logger.warning("xlsx 解析失败 %s: %s", filename, e)
                return f"[表格附件 {filename}] 解析失败: {e}"

        # CSV / TSV
        try:
            text = payload.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)
            if not rows:
                return f"[表格附件 {filename}] 空表格"
            return self._rows_to_markdown(rows, filename)
        except Exception as e:
            logger.warning("CSV 解析失败 %s: %s", filename, e)
            return f"[表格附件 {filename}] 解析失败: {e}"

    def _extract_xlsx(self, payload: bytes, filename: str) -> str:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        parts: List[str] = [f"[表格附件 {filename}]"]
        for sheet in wb.worksheets:
            parts.append(f"## 工作表: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                parts.append(self._rows_to_markdown([list(r) for r in rows], filename, header=False))
        wb.close()
        return "\n".join(parts)

    def _rows_to_markdown(
        self, rows: List[List[Any]], filename: str, header: bool = True
    ) -> str:
        if not rows:
            return f"[表格附件 {filename}] 空表格"
        # 截断过宽/过长表格，避免撑爆 prompt
        max_cols = 8
        max_rows = 50
        truncated = rows[:max_rows]
        cols = max(len(r) for r in truncated)
        cols = min(cols, max_cols)

        def _cell(v: Any) -> str:
            if v is None:
                return ""
            return str(v).replace("|", "\\|").replace("\n", " ")

        lines = []
        if header:
            head = [_cell(rows[0][i]) if i < len(rows[0]) else "" for i in range(cols)]
            lines.append("| " + " | ".join(head) + " |")
            lines.append("| " + " | ".join(["---"] * cols) + " |")
            body = truncated[1:]
        else:
            lines.append("| " + " | ".join(["列"] * cols) + " |")
            lines.append("| " + " | ".join(["---"] * cols) + " |")
            body = truncated

        for row in body:
            cells = [_cell(row[i]) if i < len(row) else "" for i in range(cols)]
            lines.append("| " + " | ".join(cells) + " |")

        if len(rows) > max_rows:
            lines.append(f"\n(表格共 {len(rows)} 行，已截断显示前 {max_rows} 行)")
        return "\n".join(lines)


class ImageExtractor(BaseExtractor):
    """
    图片附件抽取器：OCR / 视觉理解。
    优先使用注入的 vision_callable（如云端多模态模型）；未配置时降级为占位说明。
    """

    kind = "image"

    def __init__(self, vision_callable=None):
        self._vision = vision_callable

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知图片")
        if self._vision is None:
            return (
                f"[图片附件 {filename}] 未配置 OCR/视觉模型，已跳过抽取。"
                "请在系统中配置多模态模型以启用图片理解。"
            )
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[图片附件 {filename}] 无法读取图片数据"
        try:
            import base64

            b64 = base64.b64encode(payload).decode("ascii")
            mime = attachment.get("mime", "image/png")
            description = await self._vision(b64, mime, filename)
            return f"[图片附件 {filename}]\n{description}"
        except Exception as e:
            logger.warning("图片抽取失败 %s: %s", filename, e)
            return f"[图片附件 {filename}] 抽取失败: {e}"


class AudioExtractor(BaseExtractor):
    """
    音频附件抽取器：ASR 语音转文字。
    优先使用注入的 asr_callable（如 Whisper）；未配置时降级为占位说明。
    """

    kind = "audio"

    def __init__(self, asr_callable=None):
        self._asr = asr_callable

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知音频")
        if self._asr is None:
            return (
                f"[音频附件 {filename}] 未配置 ASR 语音识别模型，已跳过抽取。"
                "请在系统中配置 ASR 模型以启用语音转文字。"
            )
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[音频附件 {filename}] 无法读取音频数据"
        try:
            transcript = await self._asr(payload, filename)
            return f"[音频附件 {filename}]\n{transcript}"
        except Exception as e:
            logger.warning("音频抽取失败 %s: %s", filename, e)
            return f"[音频附件 {filename}] 抽取失败: {e}"


class PdfExtractor(BaseExtractor):
    """
    PDF 附件抽取器：提取文本层。
    优先使用 pdfplumber/PyPDF2；未安装时降级。
    """

    kind = "pdf"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知PDF")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[PDF附件 {filename}] 无法读取数据"
        try:
            return self._extract_pdf(payload, filename)
        except ImportError:
            return f"[PDF附件 {filename}] PDF 解析需安装 pdfplumber，已跳过"
        except Exception as e:
            logger.warning("PDF 解析失败 %s: %s", filename, e)
            return f"[PDF附件 {filename}] 解析失败: {e}"

    def _extract_pdf(self, payload: bytes, filename: str) -> str:
        try:
            import pdfplumber  # type: ignore

            pages_text: List[str] = []
            with pdfplumber.open(io.BytesIO(payload)) as pdf:
                for i, page in enumerate(pdf.pages[:20]):  # 最多前 20 页
                    text = page.extract_text() or ""
                    if text:
                        pages_text.append(f"--- 第 {i + 1} 页 ---\n{text}")
            body = "\n\n".join(pages_text) if pages_text else "(无可提取文本，可能是扫描件)"
            return f"[PDF附件 {filename}]\n{body}"
        except ImportError:
            # 兜底 PyPDF2
            try:
                from PyPDF2 import PdfReader  # type: ignore

                reader = PdfReader(io.BytesIO(payload))
                pages_text = []
                for i, page in enumerate(reader.pages[:20]):
                    text = page.extract_text() or ""
                    if text:
                        pages_text.append(f"--- 第 {i + 1} 页 ---\n{text}")
                body = "\n\n".join(pages_text) if pages_text else "(无可提取文本)"
                return f"[PDF附件 {filename}]\n{body}"
            except ImportError:
                raise


class UnknownExtractor(BaseExtractor):
    """未知类型附件：记录占位"""

    kind = "unknown"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知附件")
        mime = attachment.get("mime", "unknown")
        return f"[附件 {filename}（{mime}）] 暂不支持的附件类型，已跳过"
